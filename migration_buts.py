"""
================================================================================
MIGRATION  But_D1.xlsx  ->  futsal_d1.db
================================================================================
Lit le fichier Excel de recensement des buts de D1 Futsal et construit une base
SQLite propre, prête pour l'appli Streamlit (app_buts.py).

À RELANCER après chaque mise à jour du fichier Excel :
    python migration_buts.py

Étapes :
  1. lit la feuille principale "BUT D1 " -> tous les buts (source fiable)
  2. détecte AUTOMATIQUEMENT les feuilles équipes (toute feuille autre que la
     principale, contenant une colonne origine) -> origine
  3. rattache l'origine à chaque but par (journée, minute, période, buteur)
  4. normalise les libellés d'origine (placee/placée -> "Attaque placée", etc.)
  5. calcule la situation au moment du but (Menant / Égalité / Mené)
  6. écrit le tout dans futsal_d1.db

Dépendance :  pip install openpyxl
================================================================================
"""

import sqlite3
import unicodedata
from pathlib import Path
import openpyxl

# ---------------------------------------------------------------------------
# CONFIG — adapte ces chemins si besoin
# ---------------------------------------------------------------------------
XLSX_PATH   = "But_D1.xlsx"
DB_PATH     = "futsal_d1.db"
SCHEMA_PATH = "schema_buts.sql"

# Nom de la feuille principale (attention à l'espace final dans ton fichier)
FEUILLE_PRINCIPALE = "BUT D1 "

# Les feuilles "équipe" (origines) sont DÉTECTÉES AUTOMATIQUEMENT :
# toute feuille autre que la principale et possédant une colonne origine est
# lue, quel que soit son nom ou sa casse (GOAL, NANTES, SPORTING, AVION...).
# -> pour ajouter une équipe : crée son onglet dans l'Excel, relance le script.
#    Plus besoin de modifier ce fichier.


# ---------------------------------------------------------------------------
# OUTILS
# ---------------------------------------------------------------------------
def nettoie(v):
    """Normalise une cellule texte : trim + espaces multiples réduits."""
    if v is None:
        return None
    s = str(v).strip()
    s = " ".join(s.split())
    return s if s else None


def sans_accents_min(s):
    """Minuscule sans accents, pour comparer/normaliser les libellés."""
    if s is None:
        return ""
    s = str(s).lower().strip().replace("_", " ")
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return " ".join(s.split())


def est_feuille_principale(nom):
    """Vrai si 'nom' désigne la feuille des buts (insensible casse/espaces)."""
    a = sans_accents_min(nom)
    return a == sans_accents_min(FEUILLE_PRINCIPALE) or a.startswith("but d1")


# Table de normalisation des origines. Clé = forme sans accents/minuscule.
# Valeur = libellé propre affiché dans l'appli.
NORMALISATION_ORIGINE = {
    "attaque placee":   "Attaque placée",
    "attaque rapide":   "Attaque rapide",
    "transition off":   "Transition offensive",
    "transition def":   "Transition défensive",
    "touche off":       "Touche offensive",
    "touche def":       "Touche défensive",
    "corner":           "Corner",
    "coup franc":       "Coup franc",
    "jet franc":        "Jet franc (10m)",
    "penalty":          "Penalty",
    "power play":       "Power play",
    "coup d'envoi":     "Coup d'envoi",
    "coup denvoi":      "Coup d'envoi",
}


def normalise_origine(v):
    """
    Renvoie le libellé propre d'une origine.
    - "?"  -> "Non identifié"  (vidéo indisponible, origine non identifiable)
    - vide -> None             (pas encore saisi)
    - sinon -> libellé normalisé, ou capitalisé si inconnu
    """
    s = nettoie(v)
    if s is None:
        return None
    if s == "?":
        return "Non identifié"
    cle = sans_accents_min(s)
    if cle in NORMALISATION_ORIGINE:
        return NORMALISATION_ORIGINE[cle]
    # libellé inconnu : on garde mais proprement capitalisé
    return s.capitalize()


def situation_au_but(score_marque_avant, score_encaisse_avant):
    """Situation de l'équipe qui marque, JUSTE AVANT son but."""
    if score_marque_avant is None or score_encaisse_avant is None:
        return None
    if score_marque_avant > score_encaisse_avant:
        return "Menant"
    if score_marque_avant == score_encaisse_avant:
        return "Égalité"
    return "Mené"


def cle_rattachement(journee, minute, periode, joueur):
    """Clé pour relier un but de la feuille principale à son origine."""
    return (
        journee,
        minute,
        periode,
        sans_accents_min(joueur),
    )


# ---------------------------------------------------------------------------
# LECTURE
# ---------------------------------------------------------------------------
def lire_entetes(ws):
    """Renvoie {nom_colonne_normalisé: index_colonne} à partir de la ligne 1."""
    entetes = {}
    for c in range(1, ws.max_column + 1):
        nom = sans_accents_min(ws.cell(row=1, column=c).value)
        if nom:
            entetes[nom] = c
    return entetes


def col(entetes, *noms):
    """Renvoie l'index de la 1re colonne trouvée parmi 'noms' (ou None)."""
    for n in noms:
        if n in entetes:
            return entetes[n]
    return None


def lire_buts_principale(wb):
    """Lit la feuille principale -> liste de dicts (un par but)."""
    ws = wb[FEUILLE_PRINCIPALE]
    e = lire_entetes(ws)

    c_journee   = col(e, "journee")
    c_dom       = col(e, "equipe domicile", "equipe domicile ")
    c_ext       = col(e, "equipe exterieure", "equipe exterieur")
    c_marque    = col(e, "equipe marque")
    c_encaisse  = col(e, "equipe encaisse")
    c_periode   = col(e, "periode")
    c_minute    = col(e, "minute")
    c_sda       = col(e, "score dom avant")
    c_sea       = col(e, "score ext avant")
    c_joueur    = col(e, "joueur")

    buts = []
    for r in range(2, ws.max_row + 1):
        journee = ws.cell(row=r, column=c_journee).value
        if journee is None:
            continue
        marque   = nettoie(ws.cell(row=r, column=c_marque).value)
        joueur   = nettoie(ws.cell(row=r, column=c_joueur).value)
        if not marque or not joueur:
            continue

        dom      = nettoie(ws.cell(row=r, column=c_dom).value)
        ext      = nettoie(ws.cell(row=r, column=c_ext).value)
        encaisse = nettoie(ws.cell(row=r, column=c_encaisse).value)
        periode  = ws.cell(row=r, column=c_periode).value
        minute   = ws.cell(row=r, column=c_minute).value
        sda      = ws.cell(row=r, column=c_sda).value
        sea      = ws.cell(row=r, column=c_sea).value

        # score "avant" du point de vue de l'équipe qui marque :
        # si elle joue à domicile, c'est score_dom_avant, sinon score_ext_avant
        if marque == dom:
            sma, sea_ = sda, sea
        elif marque == ext:
            sma, sea_ = sea, sda
        else:
            # nom ne matche pas dom/ext (rare) : on ne déduit pas la situation
            sma, sea_ = None, None

        buts.append({
            "journee":           int(journee),
            "equipe_domicile":   dom,
            "equipe_exterieure": ext,
            "equipe_marque":     marque,
            "equipe_encaisse":   encaisse,
            "periode":           int(periode) if periode is not None else None,
            "minute":            int(minute) if minute is not None else None,
            "score_marque_avant":   int(sma) if isinstance(sma, (int, float)) else None,
            "score_encaisse_avant": int(sea_) if isinstance(sea_, (int, float)) else None,
            "joueur":            joueur,
            "origine":           None,  # rempli ensuite via les feuilles équipes
        })
    return buts


def lire_origines(wb):
    """
    Détecte automatiquement les feuilles équipes (toute feuille autre que la
    principale, possédant une colonne origine) -> dict {clé_rattachement: origine}.
    En cas de doublon de clé, la 1re origine renseignée gagne.
    """
    origines = {}
    feuilles_lues = []
    for nom_feuille in wb.sheetnames:
        if est_feuille_principale(nom_feuille):
            continue
        ws = wb[nom_feuille]
        e = lire_entetes(ws)
        c_origine = col(e, "origine but", "origine")
        if c_origine is None:
            print(f"  · feuille '{nom_feuille}' sans colonne origine, ignorée")
            continue

        c_journee = col(e, "journee")
        c_periode = col(e, "periode")
        c_minute  = col(e, "minute")
        c_joueur  = col(e, "joueur")

        n_avant = len(origines)
        for r in range(2, ws.max_row + 1):
            journee = ws.cell(row=r, column=c_journee).value
            joueur  = nettoie(ws.cell(row=r, column=c_joueur).value)
            if journee is None or not joueur:
                continue
            periode = ws.cell(row=r, column=c_periode).value
            minute  = ws.cell(row=r, column=c_minute).value
            origine = normalise_origine(ws.cell(row=r, column=c_origine).value)
            if origine is None:
                continue
            cle = cle_rattachement(
                int(journee),
                int(minute) if minute is not None else None,
                int(periode) if periode is not None else None,
                joueur,
            )
            if cle not in origines:
                origines[cle] = origine
        feuilles_lues.append(f"{nom_feuille} (+{len(origines) - n_avant})")

    if feuilles_lues:
        print("  feuilles origines détectées : " + ", ".join(feuilles_lues))
    return origines


# ---------------------------------------------------------------------------
# ÉCRITURE
# ---------------------------------------------------------------------------
def construire_base(buts):
    db = Path(DB_PATH)
    if db.exists():
        db.unlink()

    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    schema = Path(SCHEMA_PATH).read_text(encoding="utf-8")
    cur.executescript(schema)

    for b in buts:
        b["situation"] = situation_au_but(
            b["score_marque_avant"], b["score_encaisse_avant"]
        )
        cur.execute("""
            INSERT INTO but (
                journee, equipe_domicile, equipe_exterieure,
                equipe_marque, equipe_encaisse, periode, minute,
                score_marque_avant, score_encaisse_avant, situation,
                joueur, origine
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            b["journee"], b["equipe_domicile"], b["equipe_exterieure"],
            b["equipe_marque"], b["equipe_encaisse"], b["periode"], b["minute"],
            b["score_marque_avant"], b["score_encaisse_avant"], b["situation"],
            b["joueur"], b["origine"],
        ))

    conn.commit()
    return conn


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    print("=" * 70)
    print("MIGRATION  But_D1.xlsx -> futsal_d1.db")
    print("=" * 70)

    if not Path(XLSX_PATH).exists():
        raise SystemExit(f"✗ Fichier introuvable : {XLSX_PATH}")
    if not Path(SCHEMA_PATH).exists():
        raise SystemExit(f"✗ Schéma introuvable : {SCHEMA_PATH}")

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)

    print("→ Lecture de la feuille principale...")
    buts = lire_buts_principale(wb)
    print(f"  {len(buts)} buts lus")

    print("→ Lecture des origines (feuilles équipes)...")
    origines = lire_origines(wb)
    print(f"  {len(origines)} origines rattachables")

    # rattachement
    rattaches = 0
    for b in buts:
        cle = cle_rattachement(b["journee"], b["minute"], b["periode"], b["joueur"])
        if cle in origines:
            b["origine"] = origines[cle]
            rattaches += 1
    print(f"  {rattaches} buts enrichis d'une origine")

    print("→ Construction de la base...")
    conn = construire_base(buts)
    cur = conn.cursor()

    # ---- bilan de contrôle ----
    print()
    print("BILAN")
    print("-" * 40)
    cur.execute("SELECT COUNT(*) FROM but")
    print(f"  Buts en base        : {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(DISTINCT joueur) FROM but")
    print(f"  Buteurs uniques     : {cur.fetchone()[0]}")
    cur.execute("SELECT COUNT(DISTINCT equipe_marque) FROM but")
    print(f"  Équipes             : {cur.fetchone()[0]}")
    cur.execute("SELECT MIN(journee), MAX(journee) FROM but")
    j1, j2 = cur.fetchone()
    print(f"  Journées            : {j1} à {j2}")
    cur.execute("SELECT COUNT(*) FROM but WHERE origine IS NOT NULL")
    print(f"  Buts avec origine   : {cur.fetchone()[0]}")
    print()
    print("  Couverture origine par équipe :")
    cur.execute("""SELECT equipe_marque, COUNT(*),
                          SUM(CASE WHEN origine IS NOT NULL THEN 1 ELSE 0 END)
                   FROM but GROUP BY equipe_marque ORDER BY 2 DESC""")
    for eq, tot, ori in cur.fetchall():
        pct = int(100 * ori / tot) if tot else 0
        print(f"    {eq:<28} {ori:>3}/{tot:<3} {pct:>3}%")
    print()
    print("  Top 5 buteurs :")
    cur.execute("""SELECT joueur, COUNT(*) n FROM but
                   GROUP BY joueur ORDER BY n DESC LIMIT 5""")
    for nom, n in cur.fetchall():
        print(f"    {n:>3}  {nom}")

    conn.close()
    print()
    print(f"✓ Base créée : {Path(DB_PATH).resolve()}")


if __name__ == "__main__":
    main()
